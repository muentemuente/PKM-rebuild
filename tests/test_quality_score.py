"""Tests für Q1b — Zwei-Achsen-Quality-Scoring (Readiness ⊥ Integration), read-only.

Alle Tests laufen auf einem tmp-Vault; der Live-Brain-Vault wird nie berührt.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import yaml
from pipeline.quality_score import (
    BAND_NUTZBAR,
    BAND_PRODUKTIV,
    TIER_HUB,
    TIER_INSEL,
    QualityConfig,
    load_redundancy_data,
    render_report,
    resolve_redundancy_paths,
    score_vault,
    to_jsonl_records,
    write_xlsx,
)


def _fm(slug: str, **over: Any) -> dict[str, Any]:
    """Vollständiges Frontmatter inkl. aller optionalen Felder (D3 = voll)."""
    fm: dict[str, Any] = {
        "title": slug.replace("-", " ").title(),
        "slug": slug,
        "summary": "Eine kurze Zusammenfassung des Artikels.",
        "type": "knowledge-article",
        "doc_role": ["explanation"],
        "category": "grundlagen",
        "tags": ["python", "basics"],
        "sources_docs": ["D_x"],
        "source_chunks": ["D_x-S0001"],
        "status": "draft",
        "review_status": "ai_drafted",
        "confidence": "high",
        "doc_version": "0.1.0",
        "created": "2026-06-01",
        "updated": "2026-06-01",
        "last_synthesized": "2026-06-01",
        "prompt_version": "v2",
        "keyphrases": ["named entity recognition", "tokenisierung"],
        "related": [],
        "aliases": [f"{slug}-alias"],
        "subcategory": "nlp",
        "parent_concept": "nlp-grundlagen",
        "child_concepts": ["ner"],
        "used_in": ["projekt-x"],
    }
    fm.update(over)
    return fm


def _write(vault: Path, folder: str, name: str, fm: dict[str, Any] | None, body: str) -> Path:
    p = vault / folder / f"{name}.md"
    p.parent.mkdir(parents=True, exist_ok=True)
    if fm is None:
        p.write_text(body, encoding="utf-8")
    else:
        p.write_text(f"---\n{yaml.safe_dump(fm, sort_keys=False)}---\n\n{body}", encoding="utf-8")
    return p


# Standalone-Body OHNE Wikilinks (D5 niedrig) — gute Struktur (D1-D3 hoch).
_STANDALONE_BODY = (
    "## Einleitung\n\n"
    "Dies ist ein sauberer eigenstaendiger Absatz mit ausreichend Inhalt, damit die "
    "Strukturpruefung im gruenen Bereich liegt und genug Woerter vorhanden sind, um die "
    "Mindestwortzahl der Strukturdimension sicher und komfortabel zu uebertreffen ohne "
    "dass der Absatz dabei kuenstlich oder gestreckt wirkt im Lesefluss des Artikels.\n\n"
    "## Vertiefung\n\n"
    "Ein zweiter Abschnitt mit weiterem erklaerendem Text, der die Hierarchie sauber "
    "fortsetzt und die Mindest Wortzahl komfortabel ueberschreitet ohne dabei zu lang "
    "zu werden, sodass die Strukturpruefung dieses Artikels vollstaendig im gruenen "
    "Bereich bleibt und keine Abzuege wegen Laenge oder Hierarchie entstehen.\n"
)

# Body MIT Wikilinks (D5 hoch, wenn Ziele existieren + Rueck-Links vorhanden).
_LINKED_BODY = (
    "## Einleitung\n\n"
    "Dies ist ein sauberer Absatz mit ausreichend Inhalt, damit die Strukturpruefung "
    "im gruenen Bereich liegt und genug Woerter vorhanden sind, um die Mindestwortzahl "
    "der Strukturdimension sicher und komfortabel zu uebertreffen ohne dass der Absatz "
    "dabei kuenstlich oder gestreckt wirkt im Lesefluss des gesamten Artikels.\n\n"
    "Der Artikel verweist auf [[target-a]] und [[target-b]] als verwandte Themen, die "
    "den Kontext sinnvoll erweitern und die Verknuepfbarkeit des Wissens deutlich erhoehen.\n\n"
    "## Vertiefung\n\n"
    "Ein zweiter Abschnitt mit weiterem erklaerendem Text, der die Hierarchie sauber "
    "fortsetzt und die Mindest Wortzahl komfortabel ueberschreitet ohne dabei zu lang "
    "zu werden, sodass die Strukturpruefung dieses Artikels vollstaendig im gruenen "
    "Bereich bleibt und keine Abzuege wegen Laenge oder Hierarchie entstehen.\n"
)


def _many_sections_body(n: int, filler_repeats: int = 1) -> str:
    """Body mit ``n`` H2-Sektionen (keine Hierarchie-Sprünge), ``filler_repeats``-Padding."""
    filler = (
        "Dies ist ein erklaerender Satz mit etwa zehn Woertern fuer Wortzahl. " * filler_repeats
    )
    return "\n".join(f"## Abschnitt {i}\n\n{filler.strip()}\n" for i in range(n))


def _build_rich_vault(vault: Path) -> None:
    """Vault mit gut verlinktem 'clean'-Artikel + Nachbarn (für D5/D6-Höhe)."""
    _write(vault, "01_Grundlagen", "clean", _fm("clean", related=["target-a"]), _LINKED_BODY)
    for name in ("target-a", "target-b"):
        _write(
            vault,
            "01_Grundlagen",
            name,
            _fm(name, keyphrases=["named entity recognition", "tokenisierung"]),
            f"## {name}\n\nInhalt mit genuegend Woertern fuer die Strukturpruefung dieses Doks.\n",
        )
    for name in ("linker-1", "linker-2"):
        _write(
            vault,
            "01_Grundlagen",
            name,
            _fm(name),
            f"## L\n\nDieser Artikel verweist auf [[clean]] zur Vertiefung des Themas {name}.\n",
        )


def _write_redundancy_reports(
    folder: Path, *, dup_rows: list[tuple[str, str, str]], sc_members: dict[str, float]
) -> None:
    """Schreibt redundancy_report.md + synthesis_candidates.md im Parser-Format."""
    folder.mkdir(parents=True, exist_ok=True)
    red = [
        "# Redundancy-Report",
        "",
        "| Band | Slug A | Slug B | TF-IDF | Embedding | x | y |",
        "|---|---|---|---|---|---|---|",
    ]
    for band, a, b in dup_rows:
        red.append(f"| {band} | `{a}` | `{b}` | 0.5 | 0.9 | — | — |")
    (folder / "redundancy_report.md").write_text("\n".join(red) + "\n", encoding="utf-8")

    syn = ["# Synthesis-Candidates", ""]
    if sc_members:
        members = ", ".join(f"`{s}`" for s in sc_members)
        sim = next(iter(sc_members.values()))
        syn += [f"## SC_000 — {len(sc_members)} Docs (Ø-Sim {sim:.3f}, 5 Kanten)", ""]
        syn += [f"**Mitglieder:** {members}", ""]
    (folder / "synthesis_candidates.md").write_text("\n".join(syn) + "\n", encoding="utf-8")


def _by_slug(vq: Any, slug: str) -> Any:
    return next(f for f in vq.files if f.slug == slug)


# === 1. Standalone-Doc (D5/D6 niedrig) erreicht trotzdem produktiv ============


def test_standalone_low_integration_still_produktiv(tmp_path: Path) -> None:
    """Graph-Isolation darf die Readiness (Band) nicht mehr deckeln."""
    vault = tmp_path / "vault"
    _write(vault, "01_Grundlagen", "standalone", _fm("standalone"), _STANDALONE_BODY)
    work = tmp_path / "work"
    _write_redundancy_reports(work, dup_rows=[], sc_members={})  # standalone nicht Member
    red, syn = resolve_redundancy_paths(None, work)
    redundancy = load_redundancy_data(red, syn)

    vq = score_vault(vault, redundancy=redundancy)
    s = _by_slug(vq, "standalone")
    assert s.dimensions["d1"].score >= 80
    assert s.dimensions["d2"].score == 100
    assert s.dimensions["d3"].score >= 90
    assert s.dimensions["d4"].score == 100  # keine Redundanz-Überlappung
    # D5/D6 niedrig, aber Band wird allein aus Readiness (D1-D4) bestimmt.
    assert s.dimensions["d5"].score < 40
    assert s.readiness_band == BAND_PRODUKTIV
    assert s.integration_tier == TIER_INSEL


# === 2. Heading-Sprung + untagged Fence → D1/D2 senken ========================


def test_heading_jump_and_untagged_fence_lowers_d1_d2(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    _write(vault, "01_Grundlagen", "clean", _fm("clean"), _STANDALONE_BODY)
    broken_body = (
        "# Titel\n\n"
        "### Direkt zu H3 — ein Hierarchie-Sprung von H1 auf H3.\n\n"
        "Etwas Text mit ausreichend Woertern fuer die Pruefung dieses kaputten Artikels.\n\n"
        "```\nuntagged code block ohne sprach tag\n```\n"
    )
    _write(vault, "01_Grundlagen", "broken", _fm("broken"), broken_body)

    vq = score_vault(vault, redundancy=None)
    clean = _by_slug(vq, "clean")
    broken = _by_slug(vq, "broken")
    assert broken.dimensions["d1"].score < clean.dimensions["d1"].score
    assert broken.dimensions["d2"].score < clean.dimensions["d2"].score


# === 3. Fehlende Pflichtfelder → D3 senkt, Band ≤ nutzbar =====================


def test_missing_required_fields_lowers_d3(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    partial = {"title": "T", "slug": "partial", "summary": "S", "type": "knowledge-article"}
    _write(vault, "01_Grundlagen", "partial", partial, "## A\n\nKurzer Body fuer den Test.\n")

    vq = score_vault(vault, redundancy=None)
    partial_fq = _by_slug(vq, "partial")
    assert partial_fq.dimensions["d3"].score < 70
    assert partial_fq.readiness_band in (BAND_NUTZBAR, "nacharbeit")


# === 4. Zwei near-dup-Files (gefütterter Report) → D4 beide niedrig ===========


def test_near_dup_pair_lowers_d4(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    body = "## A\n\nGenug Inhalt fuer den near-dup-Test mit ausreichender Wortzahl hier.\n"
    _write(vault, "01_Grundlagen", "dup-a", _fm("dup-a"), body)
    _write(vault, "01_Grundlagen", "dup-b", _fm("dup-b"), body)
    work = tmp_path / "work"
    _write_redundancy_reports(work, dup_rows=[("near-dup", "dup-a", "dup-b")], sc_members={})
    red, syn = resolve_redundancy_paths(None, work)
    redundancy = load_redundancy_data(red, syn)

    vq = score_vault(vault, redundancy=redundancy)
    assert _by_slug(vq, "dup-a").dimensions["d4"].score == 25  # near-dup-Band-Score
    assert _by_slug(vq, "dup-b").dimensions["d4"].score == 25


# === 5. Dangling-Wikilink → D5 senkt, Penalty greift ==========================


def test_dangling_wikilink_lowers_d5(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    _write(vault, "01_Grundlagen", "linked", _fm("linked"), _LINKED_BODY)
    _write(
        vault, "01_Grundlagen", "target-a", _fm("target-a"), "## A\n\nInhalt A hier vorhanden.\n"
    )
    _write(
        vault, "01_Grundlagen", "target-b", _fm("target-b"), "## B\n\nInhalt B hier vorhanden.\n"
    )
    dangling_body = "## X\n\nVerweist auf [[gibt-es-nicht-xyz]] das nicht existiert im Vault.\n"
    _write(vault, "01_Grundlagen", "dangler", _fm("dangler"), dangling_body)

    vq = score_vault(vault, redundancy=None)
    dangler = _by_slug(vq, "dangler")
    linked = _by_slug(vq, "linked")
    assert dangler.dimensions["d5"].score < linked.dimensions["d5"].score
    assert any("Dangling" in e for e in dangler.dimensions["d5"].evidence)


# === 6. Hub-Kandidat: D5/D6 hoch → integration_tier=hub, Band davon unberührt =


def test_hub_candidate_tier_without_band_effect(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    _build_rich_vault(vault)
    work = tmp_path / "work"
    _write_redundancy_reports(
        work, dup_rows=[], sc_members={"clean": 0.85, "target-a": 0.85, "target-b": 0.85}
    )
    red, syn = resolve_redundancy_paths(None, work)
    redundancy = load_redundancy_data(red, syn)

    vq = score_vault(vault, redundancy=redundancy)
    clean = _by_slug(vq, "clean")
    assert clean.dimensions["d5"].score >= 50
    d6 = clean.dimensions["d6"].score
    assert d6 is not None
    assert d6 >= 60
    assert clean.integration_tier == TIER_HUB
    # Band kommt aus Readiness (D1-D4), nicht aus Integration.
    assert clean.readiness_band in (BAND_PRODUKTIV, BAND_NUTZBAR)
    assert clean in vq.high_value_targets()


# === 7. Achsen-Orthogonalität: gleiche D1-D4, andere D5/D6 → Band gleich ======


def test_axes_orthogonality(tmp_path: Path) -> None:
    """Zwei strukturell identische Files, nur Verlinkung verschieden → gleiches Band."""
    vault = tmp_path / "vault"
    # hub-x: drei aufgelöste Out-Links + zwei In-Links.
    body_x = (
        "## Einleitung\n\nText mit genug Woertern fuer die Strukturpruefung dieses Artikels "
        "der die Mindestwortzahl klar uebertrifft und sauber bleibt im Lesefluss komplett.\n\n"
        "Siehe [[a]], [[b]] und [[c]] fuer weitere verwandte Aspekte des Themas hier.\n\n"
        "## Vertiefung\n\nZweiter Abschnitt mit weiterem Text der die Hierarchie sauber "
        "fortsetzt und genug Inhalt fuer ein gruenes Strukturergebnis dieses Doks liefert.\n"
    )
    body_y = (
        "## Einleitung\n\nText mit genug Woertern fuer die Strukturpruefung dieses Artikels "
        "der die Mindestwortzahl klar uebertrifft und sauber bleibt im Lesefluss komplett.\n\n"
        "Ein eigenstaendiger Absatz ohne jegliche Verweise auf andere Notizen im Vault hier.\n\n"
        "## Vertiefung\n\nZweiter Abschnitt mit weiterem Text der die Hierarchie sauber "
        "fortsetzt und genug Inhalt fuer ein gruenes Strukturergebnis dieses Doks liefert.\n"
    )
    _write(vault, "01_Grundlagen", "hub-x", _fm("hub-x"), body_x)
    _write(vault, "01_Grundlagen", "iso-y", _fm("iso-y"), body_y)
    for t in ("a", "b", "c"):
        _write(
            vault, "01_Grundlagen", t, _fm(t), f"## {t}\n\nInhalt {t} mit genug Woertern hier.\n"
        )
    for p in ("p", "q"):
        _write(
            vault, "01_Grundlagen", p, _fm(p), f"## {p}\n\nVerweist auf [[hub-x]] zum Thema {p}.\n"
        )

    vq = score_vault(vault, redundancy=None)  # D4/D6 n/a → Achse B = D5 allein
    x = _by_slug(vq, "hub-x")
    y = _by_slug(vq, "iso-y")
    assert x.readiness_band == y.readiness_band  # gleiche Readiness → gleiches Band
    assert x.integration_tier != y.integration_tier  # andere Verlinkung → anderes Tertil
    assert x.dimensions["d5"].score > y.dimensions["d5"].score


# === 8. D2 typ-bewusst: compact-reference vs. gedanke =========================


def test_d2_type_aware_section_cap(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    _write(
        vault,
        "01_Grundlagen",
        "ref",
        _fm("ref", type="compact-reference"),
        _many_sections_body(40),
    )
    _write(
        vault,
        "01_Grundlagen",
        "gedanke-40",
        _fm("gedanke-40", type="gedanke"),
        _many_sections_body(40),
    )

    vq = score_vault(vault, redundancy=None)
    ref_d2 = _by_slug(vq, "ref").dimensions["d2"].score
    ged_d2 = _by_slug(vq, "gedanke-40").dimensions["d2"].score
    assert ref_d2 >= 90  # compact-reference: 40 < Cap 45 → kein Sektions-Penalty
    assert ged_d2 < ref_d2  # gedanke: Cap 8 greift → gestraft
    assert ged_d2 < 70


# === 9. Längen-Softening: langes vs. kurzes Doc mit vielen Sektionen ==========


def test_d2_length_softening(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    # langes Doc: 30 Sektionen, viel Text → effektiver Cap durch Wortzahl gelockert.
    _write(vault, "01_Grundlagen", "long", _fm("long"), _many_sections_body(30, filler_repeats=14))
    # kurzes Doc: 30 Sektionen, wenig Text → Cap 14 greift → gestraft.
    _write(vault, "01_Grundlagen", "short", _fm("short"), _many_sections_body(30, filler_repeats=1))

    vq = score_vault(vault, redundancy=None)
    long_d2 = _by_slug(vq, "long").dimensions["d2"].score
    short_d2 = _by_slug(vq, "short").dimensions["d2"].score
    assert long_d2 > 0  # nicht allein wegen Sektionszahl auf 0
    assert long_d2 > short_d2  # Längen-Softening wirkt nur beim langen Doc


# === 10. Idempotenz: zweimal scoren → identischer Hash ========================


def test_idempotent_score_hash(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    _build_rich_vault(vault)
    work = tmp_path / "work"
    _write_redundancy_reports(work, dup_rows=[], sc_members={"clean": 0.8})
    red, syn = resolve_redundancy_paths(None, work)
    redundancy = load_redundancy_data(red, syn)

    a = score_vault(vault, redundancy=redundancy)
    b = score_vault(vault, redundancy=redundancy)
    assert a.score_hash() == b.score_hash()
    assert a.score_map() == b.score_map()


# === 11. n/a-Pfad: kein Redundanz-Report → D4/D6 n/a, Achsen reskaliert =======


def test_na_path_rescales_axes(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    _build_rich_vault(vault)

    vq = score_vault(vault, redundancy=None)  # kein Report
    clean = _by_slug(vq, "clean")
    assert clean.dimensions["d4"].na
    assert clean.dimensions["d6"].na
    # Readiness nur über D1-D3, Integration nur über D5 — kein Crash, beide 0-100.
    assert 0 <= clean.readiness_composite <= 100
    assert 0 <= clean.integration_index <= 100
    assert vq.sources_active["redundancy"] is False


# === 12. (Bonus) xlsx valide + _index exkludiert ==============================


def test_xlsx_and_index_exclusion(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    _build_rich_vault(vault)
    _write(vault, "01_Grundlagen", "_index", None, "# Index\n\nNicht bewerten.\n")

    vq = score_vault(vault, redundancy=None)
    assert all(f.slug != "_index" for f in vq.files)
    assert all(not f.relpath.endswith("_index.md") for f in vq.files)

    out = write_xlsx(vq, tmp_path / "out" / "scores.xlsx")
    assert out.is_file()
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None
    assert ws.max_row == len(vq.files) + 1  # Header + Datenzeilen


# === 13. Report + JSONL Smoke (neue Achsen-Felder) ============================


def test_report_and_jsonl_render(tmp_path: Path) -> None:
    vault = tmp_path / "vault"
    _build_rich_vault(vault)
    vq = score_vault(vault, redundancy=None, cfg=QualityConfig())
    report = render_report(vq, QualityConfig(), top_n=5)
    assert "Readiness-Band-Verteilung" in report
    assert "Leverage-Quadrant" in report
    assert "High-Value" in report
    recs = to_jsonl_records(vq)
    assert len(recs) == len(vq.files)
    assert all("readiness_composite" in r and "integration_tier" in r for r in recs)
    assert all("composite" not in r for r in recs)  # altes 6-dim-Composite entfernt
