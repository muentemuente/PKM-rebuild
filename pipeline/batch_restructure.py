"""WP3c-6 — Batch-restructure + openpyxl-Review-Sheet + review-ingest.

Skaliert das typ-bewusste restructure (WP3c-4) auf mehrere Files und stellt dem
Owner ein Entscheidungs-Sheet bereit. **Die gesamte Kette ist review-Tier**: es
entstehen ausschließlich Drafts (`pkm-pipeline/drafts/_wp3c6/`) — **kein Vault-Write,
kein D4**. Die Promotion läuft separat über WP3c-5 (`pkm promote`, Owner-Gate).

Drei Bausteine:

1. :func:`run_batch_restructure` — pro File ein Draft; ein Fail (Timeout/Parse)
   stoppt den Batch **nicht**, sondern landet in der Fehl-Liste (needs_human).
2. :func:`write_review_sheet` — `.xlsx` mit einer Zeile/Draft, Owner-Decision-Dropdown
   (accept/reject/edit) und Hervorhebung von low-confidence / `reclassified` /
   `new`-ohne-Pflichtfelder.
3. :func:`ingest_review_sheet` — liest die Entscheidungen: ``accept`` → Draft auf
   ``review_status: human_reviewed`` (nur Frontmatter); ``reject`` → ``archive/``;
   ``edit`` → bleibt liegen, geflaggt. Kein Vault-Write.
"""

from __future__ import annotations

import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from pipeline import _paths
from pipeline.config import QwenConfig
from pipeline.restructure import RestructureError, restructure_file
from pipeline.schemas import FrontmatterDraft
from pipeline.vault_audit import parse_frontmatter, split_frontmatter

#: Spalten des Review-Sheets (Reihenfolge = Sheet-Spalten A…J).
SHEET_COLUMNS = (
    "slug",
    "type",
    "type_source",
    "restructure_action",
    "confidence",
    "promote_mode",
    "genre_shift_flag",
    "runtime_s",
    "draft_path",
    "owner_decision",
)
_DECISION_COL = "J"  # owner_decision (1-basiert: 10. Spalte)
_OWNER_DECISIONS = ("accept", "reject", "edit")

_FILL_LOW = PatternFill("solid", fgColor="FFF2CC")  # low confidence (gelb)
_FILL_RECLASS = PatternFill("solid", fgColor="FCE4D6")  # reclassified (orange)
_FILL_INCOMPLETE = PatternFill("solid", fgColor="F8CBAD")  # new-ohne-Pflichtfelder (rot)


@dataclass(frozen=True)
class BatchRow:
    """Eine Draft-Zeile fürs Review-Sheet."""

    slug: str
    type: str
    type_source: str
    restructure_action: str
    confidence: str
    promote_mode: str  # "update" | "new"
    genre_shift_flag: bool
    runtime_s: float
    draft_path: Path
    new_incomplete: bool


@dataclass
class BatchResult:
    """Ergebnis eines Batch-Laufs."""

    rows: list[BatchRow] = field(default_factory=list)
    failures: list[tuple[Path, str]] = field(default_factory=list)


@dataclass
class IngestResult:
    """Ergebnis eines review-ingest-Laufs."""

    ready: list[Path] = field(default_factory=list)  # promotion-bereit (human_reviewed)
    edits: list[str] = field(default_factory=list)  # geflaggt (bleibt liegen)
    rejected: list[Path] = field(default_factory=list)  # nach archive/ verschoben


# === Hilfen (read-only) =======================================================


def _draft_frontmatter(draft_path: Path) -> dict[str, Any] | None:
    fm_text, _, _ = split_frontmatter(draft_path.read_text(encoding="utf-8"))
    if not fm_text:
        return None
    data, _ = parse_frontmatter(fm_text)
    return data


def _draft_complete(draft_path: Path) -> bool:
    """True, wenn das Draft-Frontmatter ein vollständiges Vault-Frontmatter ist."""
    data = _draft_frontmatter(draft_path)
    if not data:
        return False
    try:
        FrontmatterDraft.model_validate(data)
        return True
    except Exception:
        return False


def _slug_in_vault(vault_dir: Path, slug: str) -> bool:
    """Read-only: existiert ``<slug>.md`` (Nicht-Index) irgendwo im Vault?"""
    return any(p.name != "_index.md" for p in vault_dir.rglob(f"{slug}.md"))


# === Batch-Runner =============================================================


def run_batch_restructure(
    files: list[Path],
    *,
    client: Any,
    qwen: QwenConfig,
    vault_dir: Path,
    out_dir: Path | None = None,
    prompts_dir: Path | None = None,
    timestamp: str | None = None,
) -> BatchResult:
    """Re-strukturiert ``files`` zu Drafts; ein Fail stoppt den Batch nicht.

    Args:
        files: explizite File-Liste (opt-in, **kein** impliziter All-Vault-Lauf).
        client: injizierter Qwen-Client.
        qwen: Qwen-Konfiguration.
        vault_dir: Live-Vault (read-only) — nur für den ``promote_mode``-Check.
        out_dir: Draft-Zielordner (Default ``_paths.DRAFTS / "_wp3c6"``).
        prompts_dir / timestamp: an :func:`restructure_file` durchgereicht.
    """
    out = out_dir if out_dir is not None else (_paths.DRAFTS / "_wp3c6")
    result = BatchResult()
    for f in files:
        t0 = time.monotonic()
        try:
            draft = restructure_file(
                f,
                client=client,
                qwen=qwen,
                out_dir=out,
                prompts_dir=prompts_dir,
                timestamp=timestamp,
            )
        except RestructureError as exc:
            result.failures.append((f, str(exc)[:200]))
            continue
        runtime = time.monotonic() - t0
        promote_mode = "update" if _slug_in_vault(vault_dir, draft.slug) else "new"
        new_incomplete = promote_mode == "new" and not _draft_complete(draft.draft_path)
        result.rows.append(
            BatchRow(
                slug=draft.slug,
                type=draft.type,
                type_source=draft.type_source,
                restructure_action=draft.restructure_action,
                confidence=draft.confidence,
                promote_mode=promote_mode,
                # rewrite (nicht passthrough) → Owner prüft auf Genre-Shift.
                genre_shift_flag=draft.restructure_action == "rewrite",
                runtime_s=round(runtime, 1),
                draft_path=draft.draft_path,
                new_incomplete=new_incomplete,
            )
        )
    if result.failures:
        _write_needs_human(out, result.failures)
    return result


def _write_needs_human(out_dir: Path, failures: list[tuple[Path, str]]) -> None:
    """Schreibt die Fehl-Liste als `needs_human.txt` in den Out-Ordner."""
    out_dir.mkdir(parents=True, exist_ok=True)
    lines = [f"{src}\t{reason}" for src, reason in failures]
    (out_dir / "needs_human.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")


# === Review-Sheet (openpyxl) ==================================================


def write_review_sheet(result: BatchResult, sheet_path: Path) -> Path:
    """Schreibt ein `.xlsx`-Review-Sheet (eine Zeile/Draft, Owner-Decision-Dropdown)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "restructure-review"
    ws.append(list(SHEET_COLUMNS))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in result.rows:
        ws.append(
            [
                r.slug,
                r.type,
                r.type_source,
                r.restructure_action,
                r.confidence,
                r.promote_mode,
                "ja" if r.genre_shift_flag else "nein",
                r.runtime_s,
                str(r.draft_path),
                "",  # owner_decision (vom Owner zu füllen)
            ]
        )
        row_idx = ws.max_row
        # Hervorhebung (direkt, deterministisch — testbar via cell.fill).
        if r.confidence == "low":
            ws.cell(row=row_idx, column=5).fill = _FILL_LOW
        if r.type_source == "reclassified":
            ws.cell(row=row_idx, column=3).fill = _FILL_RECLASS
        if r.new_incomplete:
            ws.cell(row=row_idx, column=6).fill = _FILL_INCOMPLETE

    # Owner-Decision-Dropdown (Datenvalidierung) über die Datenzeilen.
    last = len(result.rows) + 1
    dv = DataValidation(type="list", formula1=f'"{",".join(_OWNER_DECISIONS)}"', allow_blank=True)
    ws.add_data_validation(dv)
    if last >= 2:
        dv.add(f"{_DECISION_COL}2:{_DECISION_COL}{last}")

    sheet_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(sheet_path)
    return sheet_path


# === review-ingest ============================================================


def _set_review_status(draft_path: Path, status: str) -> None:
    """Setzt ``review_status`` im Draft-Frontmatter (nur Frontmatter, kein Vault-Write)."""
    fm_text, body, _ = split_frontmatter(draft_path.read_text(encoding="utf-8"))
    data, _ = parse_frontmatter(fm_text) if fm_text else (None, None)
    if not data:
        raise ValueError(f"Draft-Frontmatter nicht parsebar: {draft_path}")
    data["review_status"] = status
    dumped = yaml.safe_dump(data, sort_keys=False, allow_unicode=True)
    draft_path.write_text(f"---\n{dumped}---\n\n{body.strip()}\n", encoding="utf-8")


def _archive_draft(draft_path: Path, archive_dir: Path) -> Path:
    dest_dir = archive_dir / "rejected_drafts"
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / draft_path.name
    draft_path.rename(dest)
    return dest


def ingest_review_sheet(
    sheet_path: Path,
    *,
    archive_dir: Path | None = None,
) -> IngestResult:
    """Liest die Owner-Entscheidungen aus dem Sheet und wendet sie an (kein Vault-Write).

    - ``accept`` → ``review_status: human_reviewed`` (bei ``promote_mode: new`` ohne
      vollständige Pflichtfelder → stattdessen ``edit``-Flag).
    - ``reject`` → Draft nach ``archive/rejected_drafts/``.
    - ``edit`` / leer → bleibt liegen (geflaggt bei ``edit``).
    """
    archive = archive_dir if archive_dir is not None else _paths.ARCHIVE
    wb = load_workbook(sheet_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    result = IngestResult()

    for row in ws.iter_rows(min_row=2, values_only=True):
        slug = str(row[idx["slug"]])
        decision = str(row[idx["owner_decision"]] or "").strip().lower()
        promote_mode = str(row[idx["promote_mode"]])
        draft_path = Path(str(row[idx["draft_path"]]))
        if not draft_path.exists():
            continue

        if decision == "accept":
            if promote_mode == "new" and not _draft_complete(draft_path):
                result.edits.append(slug)  # unvollständig → nicht promotierbar, edit
            else:
                _set_review_status(draft_path, "human_reviewed")
                result.ready.append(draft_path)
        elif decision == "reject":
            result.rejected.append(_archive_draft(draft_path, archive))
        elif decision == "edit":
            result.edits.append(slug)
        # leer/unbekannt → überspringen (bleibt liegen)

    return result
