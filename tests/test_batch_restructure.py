"""Tests für WP3c-6 — Batch-restructure + Review-Sheet + review-ingest.

LLM + Vault gemockt. Die ganze Kette ist review-Tier: kein Vault-Write.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import httpx
import yaml
from openai import APITimeoutError
from openpyxl import load_workbook
from pipeline.batch_restructure import (
    BatchResult,
    BatchRow,
    ingest_review_sheet,
    run_batch_restructure,
    write_review_sheet,
)
from pipeline.config import (
    QwenConfig,
    QwenMaxTokensConfig,
    QwenRestructureConfig,
    QwenTemperatureConfig,
)

_TS = "2026-06-22T00:00:00+00:00"
_STAGE3 = "# Titel\n\nNeu strukturierter Body."


def _resp(content: str) -> MagicMock:
    ch = MagicMock()
    ch.message.content = content
    ch.finish_reason = "stop"
    r = MagicMock()
    r.choices = [ch]
    return r


def _stage4(fm: dict[str, Any]) -> str:
    return f"```json\n{json.dumps(fm)}\n```"


def _qwen_cfg() -> QwenConfig:
    return QwenConfig(
        endpoint="http://localhost:1234/v1",
        model="qwen/qwen3.6-27b",
        context_window=49152,
        prompt_version="v1",
        json_mode=False,
        max_retries=0,
        retry_backoff_seconds=0,
        timeout_seconds=1200,
        temperature=QwenTemperatureConfig(stage3_synthesis=0.4, stage4_frontmatter=0.1),
        max_tokens=QwenMaxTokensConfig(stage3=16000, stage4=10000),
        restructure=QwenRestructureConfig(
            prompt_version="v2",
            reasoning_effort="none",
            temperature=0.7,
            top_p=0.8,
            presence_penalty=1.5,
            max_tokens_stage3=4000,
            max_tokens_stage4=2000,
        ),
    )


def _prompts_dir(tmp_path: Path) -> Path:
    v2 = tmp_path / "prompts" / "v2"
    v2.mkdir(parents=True)
    (v2 / "stage3_synthesis.md").write_text("S3.\n", encoding="utf-8")
    (v2 / "stage4_frontmatter_json.md").write_text("S4.\n", encoding="utf-8")
    return tmp_path / "prompts"


def _src(tmp_path: Path, slug: str, *, headings: bool = False) -> Path:
    body = "## A\n\nInhalt.\n" if headings else "Fließtext ohne Headings.\n"
    p = tmp_path / "src" / f"{slug}.md"
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(f"---\nslug: {slug}\ntype: compact-reference\n---\n\n{body}", encoding="utf-8")
    return p


# === Batch-Runner =============================================================


def test_batch_isolates_single_failure(tmp_path: Path) -> None:
    """Ein Timeout-Fail stoppt den Batch nicht; restliche Drafts entstehen, Fehl-Liste stimmt."""
    ok1 = _src(tmp_path, "ok-eins")
    bad = _src(tmp_path, "kaputt")
    ok2 = _src(tmp_path, "ok-zwei")
    out = tmp_path / "drafts"
    vault = tmp_path / "vault"
    vault.mkdir()

    # 1 gutes File = 2 Calls (Nicht-Article → kein classify; rewrite: stage3+stage4).
    def _make_client_for(slug: str) -> MagicMock:
        c = MagicMock()
        c.chat.completions.create.side_effect = [
            _resp(f"```markdown\n{_STAGE3}\n```"),
            _resp(_stage4({"title": slug, "confidence": "high"})),
        ]
        return c

    # Reihenfolge: ok-eins (2 ok), kaputt (timeout auf 1. Call), ok-zwei (2 ok).
    client = MagicMock()
    client.chat.completions.create.side_effect = [
        _resp(f"```markdown\n{_STAGE3}\n```"),
        _resp(_stage4({"title": "ok-eins", "confidence": "high"})),
        APITimeoutError(request=httpx.Request("POST", "http://x")),
        _resp(f"```markdown\n{_STAGE3}\n```"),
        _resp(_stage4({"title": "ok-zwei", "confidence": "high"})),
    ]

    result = run_batch_restructure(
        [ok1, bad, ok2],
        client=client,
        qwen=_qwen_cfg(),
        vault_dir=vault,
        out_dir=out,
        prompts_dir=_prompts_dir(tmp_path),
        timestamp=_TS,
    )
    assert {r.slug for r in result.rows} == {"ok-eins", "ok-zwei"}
    assert len(result.failures) == 1
    assert result.failures[0][0] == bad
    assert (out / "needs_human.txt").exists()
    assert not (out / "kaputt.md").exists()  # kein Draft für den Fail


def test_promote_mode_update_vs_new(tmp_path: Path) -> None:
    """promote_mode aus Slug-Existenz im Live-Vault: update vs new."""
    src_u = _src(tmp_path, "bestand")
    src_n = _src(tmp_path, "ganz-neu")
    vault = tmp_path / "vault"
    (vault / "01_Grundlagen").mkdir(parents=True)
    (vault / "01_Grundlagen" / "bestand.md").write_text(
        "---\nslug: bestand\n---\n\nx\n", encoding="utf-8"
    )

    client = MagicMock()
    client.chat.completions.create.side_effect = [
        _resp(f"```markdown\n{_STAGE3}\n```"),
        _resp(_stage4({"title": "bestand", "confidence": "high"})),
        _resp(f"```markdown\n{_STAGE3}\n```"),
        _resp(_stage4({"title": "ganz-neu", "confidence": "low"})),
    ]
    result = run_batch_restructure(
        [src_u, src_n],
        client=client,
        qwen=_qwen_cfg(),
        vault_dir=vault,
        out_dir=tmp_path / "drafts",
        prompts_dir=_prompts_dir(tmp_path),
        timestamp=_TS,
    )
    by_slug = {r.slug: r for r in result.rows}
    assert by_slug["bestand"].promote_mode == "update"
    assert by_slug["ganz-neu"].promote_mode == "new"
    # ganz-neu: new + unvollständiges Draft-Frontmatter → new_incomplete
    assert by_slug["ganz-neu"].new_incomplete is True


# === Review-Sheet =============================================================


def _row(**over: Any) -> BatchRow:
    base: dict[str, Any] = dict(
        slug="s",
        type="compact-reference",
        type_source="frontmatter",
        restructure_action="passthrough",
        confidence="high",
        promote_mode="update",
        genre_shift_flag=False,
        runtime_s=1.0,
        draft_path=Path("/d/s.md"),
        new_incomplete=False,
    )
    base.update(over)
    return BatchRow(**base)


def test_sheet_columns_dropdown_and_highlights(tmp_path: Path) -> None:
    """Sheet enthält alle Spalten + owner_decision-Dropdown + Hervorhebungen."""
    result = BatchResult(
        rows=[
            _row(slug="a", confidence="low"),
            _row(
                slug="b",
                type_source="reclassified",
                restructure_action="rewrite",
                genre_shift_flag=True,
            ),
            _row(slug="c", promote_mode="new", new_incomplete=True),
        ]
    )
    path = write_review_sheet(result, tmp_path / "sheet.xlsx")
    wb = load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    assert headers == [
        "slug",
        "type",
        "type_source",
        "restructure_action",
        "confidence",
        "promote_mode",
        "genre_shift_flag",
        "runtime_s",
        "draft_path",
        "owner_decision",
    ]
    # Dropdown vorhanden auf J2:J4
    assert any(
        "accept,reject,edit" in (dv.formula1 or "") for dv in ws.data_validations.dataValidation
    )
    # Hervorhebungen (Fill gesetzt)
    assert ws.cell(row=2, column=5).fill.fgColor.rgb.endswith("FFF2CC")  # low confidence
    assert ws.cell(row=3, column=3).fill.fgColor.rgb.endswith("FCE4D6")  # reclassified
    assert ws.cell(row=4, column=6).fill.fgColor.rgb.endswith("F8CBAD")  # new_incomplete


# === review-ingest ============================================================


def _draft(tmp_path: Path, slug: str, *, complete: bool, review: str = "ai_drafted") -> Path:
    fm: dict[str, Any] = {"slug": slug, "type": "knowledge-article", "review_status": review}
    if complete:
        fm = {
            "title": "T",
            "slug": slug,
            "summary": "S",
            "type": "knowledge-article",
            "doc_role": ["explanation"],
            "category": "grundlagen",
            "tags": ["t"],
            "sources_docs": ["D_x"],
            "source_chunks": ["D_x-S0"],
            "status": "draft",
            "review_status": review,
            "confidence": "high",
            "doc_version": "0.1.0",
            "created": "2026-06-01",
            "updated": "2026-06-01",
            "last_synthesized": "2026-06-01",
            "prompt_version": "v2",
        }
    p = tmp_path / "drafts" / f"{slug}.md"
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(
        f"---\n{yaml.safe_dump(fm, sort_keys=False)}---\n\n# T\n\nBody.\n", encoding="utf-8"
    )
    return p


def _sheet_with(tmp_path: Path, entries: list[tuple[Path, str, str]]) -> Path:
    """entries: (draft_path, promote_mode, owner_decision) → BatchResult → Sheet."""
    rows = [_row(slug=p.stem, draft_path=p, promote_mode=mode) for p, mode, _ in entries]
    path = write_review_sheet(BatchResult(rows=rows), tmp_path / "sheet.xlsx")
    # owner_decision (Spalte J) nachtragen
    wb = load_workbook(path)
    ws = wb.active
    for i, (_, _, decision) in enumerate(entries, start=2):
        ws.cell(row=i, column=10, value=decision)
    wb.save(path)
    return path


def test_ingest_accept_reject_edit(tmp_path: Path) -> None:
    """accept(update)→human_reviewed; accept(new,unvollständig)→edit; reject→archive."""
    d_update = _draft(tmp_path, "update-doc", complete=True)
    d_new_bad = _draft(tmp_path, "new-unvollstaendig", complete=False)
    d_reject = _draft(tmp_path, "weg-damit", complete=True)
    archive = tmp_path / "archive"

    sheet = _sheet_with(
        tmp_path,
        [
            (d_update, "update", "accept"),
            (d_new_bad, "new", "accept"),
            (d_reject, "update", "reject"),
        ],
    )
    result = ingest_review_sheet(sheet, archive_dir=archive)

    assert result.ready == [d_update]
    assert result.edits == ["new-unvollstaendig"]  # new + unvollständig → edit statt human_reviewed
    assert len(result.rejected) == 1

    # accept(update): review_status im Draft gesetzt (nur Frontmatter)
    fm = yaml.safe_load(d_update.read_text(encoding="utf-8").split("---\n")[1])
    assert fm["review_status"] == "human_reviewed"
    # reject: Draft verschoben nach archive/
    assert not d_reject.exists()
    assert (archive / "rejected_drafts" / "weg-damit.md").exists()
    # new-unvollständig: NICHT human_reviewed, bleibt liegen
    assert d_new_bad.exists()
    fm_bad = yaml.safe_load(d_new_bad.read_text(encoding="utf-8").split("---\n")[1])
    assert fm_bad["review_status"] == "ai_drafted"


def test_ingest_no_vault_write(tmp_path: Path) -> None:
    """review-ingest berührt keinen Vault — nur Draft-Frontmatter/Archiv."""
    vault = tmp_path / "vault"
    vault.mkdir()
    d = _draft(tmp_path, "x", complete=True)
    sheet = _sheet_with(tmp_path, [(d, "update", "accept")])
    ingest_review_sheet(sheet, archive_dir=tmp_path / "archive")
    assert list(vault.rglob("*.md")) == []
